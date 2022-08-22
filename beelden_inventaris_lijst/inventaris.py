# https://meemoo.atlassian.net/browse/OPS-1614
# to install packages
# cd ..
# make install
# source python_env/bin/activate
# cd beelden_inventaris_lijst
# python inventaris.py
#
import openpyxl
import csv
from mediahaven_api import MediahavenApi

def set_dynamic_fields(result, csv_row):
    dynamic = result['Dynamic']
    csv_row['dc_title']= dynamic.get('dc_title')
    csv_row['dc_source']= dynamic.get('dc_source')
    csv_row['dc_rechtenstatus']= dynamic.get('dc_rights_rightsHolders')
    csv_row['dc_creators']= dynamic.get('dc_creators')
    csv_row['dc_rights_credit']= dynamic.get('dc_rigths_credit')
    csv_row['dc_creators_maker']= dynamic['dc_creators'].get('Maker')
    csv_row['dc_creators_fotograaf']= dynamic['dc_creators'].get('Fotograaf')
    csv_row['dc_identifier_localid']= dynamic.get('dc_identifier_localid')
    csv_row['dc_pid']= dynamic.get('PID')
    csv_row['dc_titles_archief']= dynamic['dc_titles'].get('archief')
    csv_row['dc_titles_deelarchief']= dynamic['dc_titles'].get('deelarchief')
    return csv_row

def set_local_ids(result, csv_row):
    li = result['Dynamic']['dc_identifier_localids']    
    csv_row['li_persistente_uri_werk'] = li.get('PersistenteURI_Werk')
    csv_row['li_workpid'] = li.get('WorkPID')
    csv_row['li_inventarisnummer'] = li.get('Inventarisnummer')
    csv_row['li_afbeelding'] = li.get('Afbeelding')
    csv_row['li_objectnaam'] = li.get('Objectnaam')
    csv_row['li_persistente_uri_record'] = li.get('PersistenteURI_Record')
    return csv_row

def set_technical_fields(result, csv_row):
    technical = result['Technical'] 
    csv_row['file_size'] = technical['FileSize']
    csv_row['mimetype'] = technical['MimeType']
    csv_row['height'] = technical['Height']
    csv_row['width'] = technical['Width']
    csv_row['orientation'] = technical['ImageOrientation']
    return csv_row

def set_descriptive_fields(result, csv_row):
    descriptive = result['Descriptive']               
    csv_row['creation_date'] = descriptive.get('CreationDate')
    csv_row['author'] = descriptive['Authors'].get('Author')
    csv_row['original_filename'] = descriptive.get('OriginalFilename')
    csv_row['title'] = descriptive.get('Title')
    return csv_row


def csv_cols():
    return ['instelling', 'inventaris_nr', 'fragment_id', 'copyright', 'vervaardiger1', 'vervaardiger2', 'vervaardiger3', 'vervaardiger4', 'vervaardiger5', 'vervaardiger6', 'vervaardiger7', 'dc_title', 'dc_source', 'dc_rechtenstatus', 'dc_creators', 'dc_rights_credit', 'dc_creators_maker', 'dc_creators_fotograaf', 'dc_identifier_localid', 'dc_pid', 'dc_titles_archief', 'dc_titles_deelarchief', 'li_persistente_uri_werk', 'li_workpid', 'li_inventarisnummer', 'li_afbeelding', 'li_objectnaam', 'li_persistente_uri_record', 'file_size', 'mimetype', 'height', 'width', 'orientation', 'creation_date', 'author', 'original_filename', 'title']

def write_csv_header(writer):
    colnames = csv_cols()
    first_row = [col.capitalize() for col in colnames]
    writer.writerow(first_row)

def write_csv(writer, csv_row):
    data = []
    for col in csv_cols():
        data.append( csv_row.get(col) )

    print(f"row = {data}\n")
    writer.writerow(data)

def process_inventaris_excel(mh_api, filename, outputfile):
    SKIP_ROWS=1851  # 1 

    wb = openpyxl.load_workbook(filename, read_only=True)
    outfile = open(outputfile, 'w')
    writer = csv.writer(outfile)
    write_csv_header(writer)

    count = 0
    for ws in wb:
        for row in ws.rows:
            inventaris_nr = row[1].value
            count += 1
            if inventaris_nr == 'Inventarisnummer':
                continue

            if count < SKIP_ROWS:
                continue

            instelling = row[0].value
            csv_row = {
                'instelling': instelling,
                'inventaris_nr': inventaris_nr,
                'fragment_id': 'LOOKUP PENDING...',
                'copyright': row[2].value,
                'vervaardiger1': row[3].value,
                'vervaardiger2': row[4].value,
                'vervaardiger3': row[5].value,
                'vervaardiger4': row[6].value,
                'vervaardiger5': row[7].value,
                'vervaardiger6': row[8].value,
                'vervaardiger7': row[9].value
            }

            mh_result = mh_api.find_vkc_record(inventaris_nr)
            if mh_result:
                result = mh_api.find_fragment(mh_result['fragment_id'])

                csv_row['fragment_id'] = mh_result.get('fragment_id') 
                csv_row = set_dynamic_fields(result, csv_row) 
                csv_row = set_local_ids(result, csv_row)
                csv_row = set_technical_fields(result, csv_row)
                csv_row = set_descriptive_fields(result, csv_row)

            else:
                csv_row['fragment_id'] = 'INVENTARIS NOT FOUND IN MEDIAHAVEN'

            write_csv(writer, csv_row)
            print(
                    "processed = ", count,
                    "total rows=", ws.max_row,
                    " procent processed=",
                    float(count)/float(ws.max_row)*100,
                    "\n"
            )

    outfile.close()


def main():
    mh_api = MediahavenApi()
    process_inventaris_excel(
        mh_api,
        "OPS1614_inventarisnrs_20220804_copyright.xlsx",
        "inventaris_output_deel2.csv"
    )


if __name__ == '__main__':
    main()



